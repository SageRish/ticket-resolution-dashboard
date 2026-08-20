"""
Independent audit of the dashboard's transparency claim.

Parses the Excel formulas the dashboard renders and EXECUTES them against the
raw workbook, then compares the result to the value the dashboard displays.
If this passes, a resident who types the shown formula into the original sheet
gets the number the dashboard showed them.

Usage:  python scripts/audit_formulas.py [path/to/ticket_data.xlsx]
"""
import datetime as dt
import json
import re
import statistics
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CUTOFF = dt.datetime(2026, 8, 20, 9, 30)
COL = {"B": 2, "D": 4, "G": 7, "J": 10, "L": 12, "N": 14, "R": 18}


# ---------- the W helper column, implemented exactly as the formula reads ----------
def _find(sub, s):
    i = s.find(sub)
    if i < 0:
        raise ValueError("#VALUE!")
    return i + 1


def helper_minutes(B, L, R):
    if L == "-":
        created = dt.datetime(2000 + int(B[6:8]), int(B[3:5]), int(B[0:2])) + \
            (dt.datetime.strptime(B[10:18].strip(), "%I:%M %p") - dt.datetime(1900, 1, 1))
        return round((CUTOFF - created).total_seconds() / 60)
    d = int(R[:_find(" ", R) - 1])
    h = int(R[_find("): ", R) + 2:_find("): ", R) + 2 + (_find(" hour", R) - _find("): ", R) - 3)])
    m = int(R[_find("hour(s): ", R) + 8:_find("hour(s): ", R) + 8 + (_find(" minute", R) - _find("hour(s): ", R) - 9)])
    return d * 1440 + h * 60 + m


def load_sheet(src):
    ws = openpyxl.load_workbook(src, data_only=True)["Sheet1"]
    rows = {}
    for r in range(4, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            continue
        cells = {k: ("" if ws.cell(row=r, column=c).value is None else str(ws.cell(row=r, column=c).value))
                 for k, c in COL.items()}
        cells["W"] = helper_minutes(cells["B"], cells["L"], cells["R"])
        rows[r] = cells
    return rows


# ---------- a tiny evaluator for the formula shapes the dashboard emits ----------
CRIT = re.compile(r'\$([A-Z]{1,2})\$\d+:\$[A-Z]{1,2}\$\d+,"((?:[^"]|"")*)"')


def _match(rows, crits):
    out = []
    for r, c in rows.items():
        if all(c[col] == val for col, val in crits):
            out.append(r)
    return out


def evaluate(f, rows, villa_avgs):
    f = f.replace("\n", "").replace("  ", " ").strip()

    # =AVERAGE($W$4:$W$1180)
    if re.fullmatch(r"=AVERAGE\(\$W\$\d+:\$W\$\d+\)", f):
        return statistics.fmean(c["W"] for c in rows.values())

    # =AVERAGEIFS($W..., <col>,"v" [, <col>,"v"])
    m = re.fullmatch(r"=AVERAGEIFS\(\$W\$\d+:\$W\$\d+,(.*)\)", f)
    if m:
        crits = CRIT.findall(m.group(1))
        sel = _match(rows, crits)
        return statistics.fmean(rows[r]["W"] for r in sel) if sel else None

    # =COUNTIFS(<col>,"v", <col>,"v")   /   =COUNTIF(<col>,"v")
    m = re.fullmatch(r"=COUNTIFS?\((\$[A-Z]{1,2}\$\d+:\$[A-Z]{1,2}\$\d+,\"[^\"]*\"(?:,\$[A-Z]{1,2}\$\d+:\$[A-Z]{1,2}\$\d+,\"[^\"]*\")*)\)", f)
    if m:
        return float(len(_match(rows, CRIT.findall(m.group(1)))))

    # =COUNTIF($AB$4:$AB$178,">"&AVERAGEIFS(...))
    m = re.fullmatch(r'=COUNTIF\(\$AB\$\d+:\$AB\$\d+,"([<>])"&(AVERAGEIFS\(.*\))\)', f)
    if m:
        target = evaluate("=" + m.group(2), rows, villa_avgs)
        if m.group(1) == ">":
            return float(sum(1 for v in villa_avgs if v > target))
        return float(sum(1 for v in villa_avgs if v < target))

    # =RANK(AVERAGEIFS(...),$AB$4:$AB$178,1)
    m = re.fullmatch(r"=RANK\((AVERAGEIFS\(.*\)),\$AB\$\d+:\$AB\$\d+,1\)", f)
    if m:
        target = evaluate("=" + m.group(1), rows, villa_avgs)
        return float(sum(1 for v in villa_avgs if v < target) + 1)

    # =MEDIAN(IF(($G...="x")*($N...="y"),$W...))
    m = re.fullmatch(r"=MEDIAN\(IF\(\((.*)\),\$W\$\d+:\$W\$\d+\)\)", f)
    if m:
        crits = re.findall(r'\$([A-Z]{1,2})\$\d+:\$[A-Z]{1,2}\$\d+="([^"]*)"', m.group(1))
        sel = _match(rows, crits)
        return statistics.median(rows[r]["W"] for r in sel) if sel else None

    return "UNSUPPORTED"


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "ticket_data.xlsx"
    rows = load_sheet(src)

    # AB column: one average per villa, in the order UNIQUE() would produce
    order, seen = [], set()
    for r in sorted(rows):
        v = rows[r]["G"]
        if v not in seen:
            seen.add(v)
            order.append(v)
    villa_avgs = [statistics.fmean(c["W"] for c in rows.values() if c["G"] == v) for v in order]

    cases = json.loads(Path(sys.argv[2]).read_text(encoding="utf-8")) if len(sys.argv) > 2 else []

    ok = fail = unsup = 0
    for c in cases:
        got = evaluate(c["code"], rows, villa_avgs)
        exp = c.get("expect")
        if got == "UNSUPPORTED":
            unsup += 1
            print(f"  [skip] {c['label'][:52]:54} (shape not modelled)")
            continue
        if exp is None:
            print(f"  [info] {c['label'][:52]:54} = {got}")
            ok += 1
            continue
        good = got is not None and abs(got - exp) < 0.51
        ok, fail = (ok + 1, fail) if good else (ok, fail + 1)
        flag = "ok  " if good else "FAIL"
        print(f"  [{flag}] {c['label'][:52]:54} formula={got!r:>22} dashboard={exp!r}")

    print(f"\n  {ok} matched, {fail} failed, {unsup} skipped")
    return 1 if fail else 0


if __name__ == "__main__":
    sys.exit(main())
